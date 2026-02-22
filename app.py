from __future__ import annotations

# Веб-приложение для расчета переплаты по ипотеке.

from dataclasses import dataclass
from datetime import date
import calendar
import io
from typing import Optional

from flask import Flask, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, numbers


app = Flask(__name__)


@dataclass
class MortgageInput:
    # Входные данные пользователя для расчета.
    principal: float
    down_payment: float
    down_payment_percent: float
    extra_payment: float
    prepayment_strategy: str
    years: int
    annual_rate: float


@dataclass
class MortgageResult:
    # Результат расчета переплаты.
    overpayment: float
    total_paid: float
    monthly_payment: float
    loan_amount: float
    required_income: float
    schedule: list[dict[str, float | int]]
    total_interest: float


def calculate_mortgage(data: MortgageInput) -> MortgageResult:
    # Рассчитываем ежемесячный платеж и переплату по аннуитетной схеме.
    months = data.years * 12
    monthly_rate = data.annual_rate / 12 / 100
    # Определяем сумму первоначального взноса: фиксированная сумма или процент.
    if data.down_payment > 0:
        down_payment_value = data.down_payment
    elif data.down_payment_percent > 0:
        down_payment_value = data.principal * (data.down_payment_percent / 100)
    else:
        down_payment_value = 0.0

    loan_amount = data.principal - down_payment_value

    if months <= 0:
        raise ValueError("Срок должен быть больше нуля")
    if loan_amount <= 0:
        raise ValueError("Первоначальный взнос должен быть меньше суммы")

    if monthly_rate == 0:
        base_payment = loan_amount / months
    else:
        factor = (1 + monthly_rate) ** months
        base_payment = loan_amount * (monthly_rate * factor) / (factor - 1)

    # Банки обычно округляют ежемесячный платеж до копеек.
    base_payment = round(base_payment, 2)

    # Формируем график платежей помесячно.
    schedule: list[dict[str, float | int]] = []
    balance = loan_amount
    total_interest = 0.0

    # Стартовая дата графика: текущий день.
    start_date = date.today()

    def add_months(base_date: date, months_to_add: int) -> date:
        # Безопасно добавляем месяцы к дате, учитывая длину месяца.
        month_index = base_date.month - 1 + months_to_add
        year = base_date.year + month_index // 12
        month = month_index % 12 + 1
        last_day = calendar.monthrange(year, month)[1]
        day = min(base_date.day, last_day)
        return date(year, month, day)

    current_payment = base_payment
    month = 1

    while balance > 0 and month <= months:
        # Начисленные проценты за месяц.
        interest = round(balance * monthly_rate, 2) if monthly_rate > 0 else 0.0

        principal_payment = round(current_payment - interest, 2)
        if principal_payment < 0:
            principal_payment = 0.0

        total_principal = round(principal_payment + data.extra_payment, 2)

        # Корректируем последний платеж, чтобы закрыть остаток.
        if total_principal > balance:
            total_principal = round(balance, 2)

        monthly_payment_actual = round(interest + total_principal, 2)
        balance = round(balance - total_principal, 2)
        total_interest = round(total_interest + interest, 2)

        payment_date = add_months(start_date, month - 1)

        schedule.append(
            {
                "month": month,
                "payment": monthly_payment_actual,
                "interest": interest,
                "principal": total_principal,
                "balance": max(balance, 0.0),
                "date": payment_date.isoformat(),
            }
        )

        # Пересчитываем платеж при стратегии уменьшения платежа.
        remaining_months = months - month
        if data.prepayment_strategy == "reduce_payment" and remaining_months > 0 and balance > 0:
            if monthly_rate == 0:
                current_payment = round(balance / remaining_months, 2)
            else:
                factor = (1 + monthly_rate) ** remaining_months
                current_payment = round(balance * (monthly_rate * factor) / (factor - 1), 2)

        month += 1

    total_paid = round(sum(item["payment"] for item in schedule), 2)
    overpayment = round(total_interest, 2)

    # Оцениваем минимальный доход по правилу 30%.
    monthly_payment = schedule[0]["payment"] if schedule else base_payment
    required_income = monthly_payment / 0.3

    return MortgageResult(
        overpayment=overpayment,
        total_paid=total_paid,
        monthly_payment=monthly_payment,
        loan_amount=loan_amount,
        required_income=required_income,
        schedule=schedule,
        total_interest=total_interest,
    )


def parse_form(form) -> MortgageInput:
    # Преобразуем входные значения из формы в числа.
    # Нормализуем ввод: пустая строка считается нулем.
    def parse_float(value: str) -> float:
        cleaned = value.replace(" ", "").replace(",", ".").strip()
        if cleaned == "":
            return 0.0
        return float(cleaned)

    def parse_int(value: str) -> int:
        cleaned = value.replace(" ", "").replace(",", ".").strip()
        if cleaned == "":
            return 0
        return int(float(cleaned))

    principal = parse_float(form.get("principal", ""))
    down_payment = parse_float(form.get("down_payment", ""))
    down_payment_percent = parse_float(form.get("down_payment_percent", ""))
    extra_payment = parse_float(form.get("extra_payment", ""))
    prepayment_strategy = form.get("prepayment_strategy", "reduce_term")
    years = parse_int(form.get("years", ""))
    annual_rate = parse_float(form.get("annual_rate", ""))

    if principal <= 0:
        raise ValueError("Сумма кредита должна быть больше нуля")
    if down_payment < 0:
        raise ValueError("Первоначальный взнос не может быть отрицательным")
    if down_payment_percent < 0:
        raise ValueError("Процент первоначального взноса не может быть отрицательным")
    if down_payment_percent >= 100:
        raise ValueError("Процент первоначального взноса должен быть меньше 100")
    if down_payment > 0 and down_payment_percent > 0:
        raise ValueError("Укажите взнос либо суммой, либо процентом")
    if down_payment >= principal:
        raise ValueError("Первоначальный взнос должен быть меньше суммы")
    if extra_payment < 0:
        raise ValueError("Досрочный платеж не может быть отрицательным")
    if prepayment_strategy not in {"reduce_term", "reduce_payment"}:
        raise ValueError("Некорректная стратегия досрочного платежа")
    if years <= 0:
        raise ValueError("Срок должен быть больше нуля")
    if annual_rate < 0:
        raise ValueError("Ставка не может быть отрицательной")

    return MortgageInput(
        principal=principal,
        down_payment=down_payment,
        down_payment_percent=down_payment_percent,
        extra_payment=extra_payment,
        prepayment_strategy=prepayment_strategy,
        years=years,
        annual_rate=annual_rate,
    )


def format_money(value: float) -> str:
    # Форматируем число в денежный вид с пробелами.
    return f"{value:,.2f}".replace(",", " ")


def build_excel(schedule: list[dict[str, float | int]], result: MortgageResult) -> io.BytesIO:
    # Собираем Excel-файл с графиком платежей.
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "График платежей"

    headers = [
        "Месяц",
        "Дата",
        "Платеж",
        "Проценты",
        "Основной долг",
        "Остаток",
    ]
    sheet.append(headers)

    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in schedule:
        sheet.append(
            [
                row["month"],
                row["date"],
                row["payment"],
                row["interest"],
                row["principal"],
                row["balance"],
            ]
        )

    money_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=6):
        for cell in row:
            cell.number_format = money_format

    sheet.append([])
    sheet.append(["Итоги", "", result.total_paid, result.total_interest, result.loan_amount, ""])
    for cell in sheet[sheet.max_row]:
        cell.font = header_font

    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 16
    sheet.column_dimensions["E"].width = 18
    sheet.column_dimensions["F"].width = 16

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


@app.route("/", methods=["GET", "POST"])
def index():
    # Отображаем форму и результат расчета.
    result: Optional[MortgageResult] = None
    error: Optional[str] = None
    values = {
        "principal": "",
        "down_payment": "",
        "down_payment_percent": "",
        "extra_payment": "",
        "prepayment_strategy": "reduce_term",
        "years": "",
        "annual_rate": "",
    }

    if request.method == "POST":
        try:
            data = parse_form(request.form)
            result = calculate_mortgage(data)
            values = {
                "principal": request.form.get("principal", ""),
                "down_payment": request.form.get("down_payment", ""),
                "down_payment_percent": request.form.get("down_payment_percent", ""),
                "extra_payment": request.form.get("extra_payment", ""),
                "prepayment_strategy": request.form.get("prepayment_strategy", "reduce_term"),
                "years": request.form.get("years", ""),
                "annual_rate": request.form.get("annual_rate", ""),
            }
        except (ValueError, TypeError) as exc:
            error = str(exc)

    return render_template(
        "index.html",
        result=result,
        error=error,
        values=values,
        format_money=format_money,
    )


@app.route("/export", methods=["POST"])
def export_excel():
    # Экспортируем график платежей в Excel.
    data = parse_form(request.form)
    result = calculate_mortgage(data)
    output = build_excel(result.schedule, result)

    return send_file(
        output,
        as_attachment=True,
        download_name="mortgage_schedule.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    # Локальный запуск приложения.
    app.run(debug=True, port=5050)
